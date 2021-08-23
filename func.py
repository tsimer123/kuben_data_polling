from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook


def resolve_name_file(dir_excel, type_ivk):  # функция получения адерса файла из ИВК ВУ для парсинга

    # проверка типа входных данных ИВК ВУ П или Ц
    if type_ivk == "P":
        name_ivk = " Кубань пирамида.xlsx"
    if type_ivk == "C":
        name_ivk = " Кубань Ц.xlsx"

    current_datetime = datetime.now()
    month = current_datetime.month
    day = current_datetime.day

    # првоверка на длину месяца если меньше 2 знаков то впереди добавить 0
    if month < 10:
        month = "0" + str(month)
    else:
        month = str(month)

    # првоверка на длину дня если меньше 2 знаков то впереди добавить 0
    if day < 10:
        day = "0" + str(day)
    else:
        day = str(day)

    # собираем название файла
    file_name = str(current_datetime.year) + month + day + name_ivk

    # определеить домашню директорию
    home_dir = Path.home()
    # собираем адрес до файла
    path = Path(home_dir, dir_excel, file_name)
    path_name = str(path)

    return path_name


def open_excel(name_file):  # функция получения сырого массива из файла

    # НАДО ДОБАВИТЬ ИСКЛЮЧЕНИЯ ЕСЛИ ФАЙЛА НЕТ
    wb = load_workbook(filename=name_file, read_only=True)
    ws = wb['Лист1']
    # массив для данных из excel
    mass_excel = []
    i = 0
    for row in ws.rows:
        # добавить строку в массив
        mass_excel.append([])
        for cell in row:
            # добавить содержимое ячейки в массив
            mass_excel[i].append(cell.value)
            # cell.row = 100   можно изменить строку
        i += 1
    # закрытие документа
    wb.close()

    # вернуть массив
    return mass_excel


def first_del_empty_str(data_excel):   # удаляем первые пустые строки и заголовок

    i = 0
    while i < 1:
        if data_excel[0][0] is None:
           del data_excel[0]
        else:
            del data_excel[0]
            i = 1

    return data_excel


def del_old_pu(data_excel, start_building):   # удаляются строки с ПУ старше даты из start_building

    i = 0
    len_mass = len(data_excel)
    while i < len_mass:
        if data_excel[i][15] is None:
            del data_excel[i]
            len_mass -= 1
        else:
            if data_excel[i][15] < start_building:
                del data_excel[i]
                len_mass -= 1
            else:
                i += 1

    return data_excel


def del_alien_pu(data_excel, list_type_pu):   # удаляем чужие ПУ

    i = 0
    len_mass = len(data_excel)
    while i < len_mass:
        type_pu_in_mass = data_excel[i][11]
        # удаление строки без типа ПУ
        if data_excel[i][11] is None:
            del data_excel[i]
            len_mass -= 1

        # првоверка соответствия типа ПУ с заданными
        trigger_del = 0
        for type_pu in list_type_pu:
            if type_pu_in_mass[0:2] != type_pu:
                trigger_del += 1

        #  провека размера тригера чужих пу если он меньше размера списка типов то удаляем
        if trigger_del == len(list_type_pu):
            del data_excel[i]
            len_mass -= 1
        else:
            i += 1

    return data_excel


def off_status_pu(data_excel, str_type_off_status):   # удаление строк исключенных из стартистики сбора "Не учит."

    i = 0
    len_mass = len(data_excel)
    while i < len_mass:
        status_in_mass = data_excel[i][19]
        if status_in_mass == str_type_off_status:
            del data_excel[i]
            len_mass -= 1
        else:
            i += 1

    return data_excel


def rename_empty_feeder(data_excel):  # заполнение пустых названий фидера. данные берутся из названий ТП
    i = 0
    len_mass = len(data_excel)
    while i < len_mass:
        if data_excel[i][4] is None:
            name_tp = data_excel[i][5]
            number_name_tp = name_tp.find(' ')
            data_excel[i][4] = name_tp[0:number_name_tp]
        i += 1

    return data_excel


def del_name_feeder(data_excel, list_del_name_feeder):  # удаление лишних знаков из названия фидера

    i = 0
    len_mass = len(data_excel)
    while i < len_mass:
        name_feeder_in_mass = data_excel[i][4]
        for name_feeder in list_del_name_feeder:
            if len(name_feeder_in_mass) >= len(name_feeder):
                if name_feeder_in_mass.find(name_feeder) != -1:
                    data_excel[i][4] = name_feeder_in_mass.replace(name_feeder, '')
        i += 1

    return data_excel


def save_data_excel_in_wb(data_excel):

    wb = Workbook()
    ws = wb.active
    ws.title = 'Shets1'
    i = 0
    for cells in data_excel:
        j = 0
        for cell in cells:
            ws.cell(row=i + 1, column=j + 1).value = data_excel[i][j]
            j += 1
        i += 1

    wb.save("output.xlsx")

    return data_excel
